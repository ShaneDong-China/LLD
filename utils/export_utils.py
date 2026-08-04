# -*- coding: utf-8 -*-
"""Excel 导出：每台设备一个 sheet 的 Port_Design 格式（LLD 交付物）

服务器模板设备（type=server）自动汇总到一张「服务器」sheet，
用「设备」列区分，不再每台服务器一个 sheet。
"""
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .project_utils import get_project, get_project_devices, get_device_ports, port_display_name
from .connection_utils import (get_connections, get_connection_by_port,
                               conn_port_a_display, conn_port_b_display)

HEADERS = ['设备', '位置', '接口', '接口状态', '聚合组', '聚合模式', '接口模式', 'VLAN', '对端设备', '对端接口', '连接描述', '备注']
WIDTHS = [16, 14, 24, 10, 12, 11, 11, 14, 16, 24, 34, 16]
SERVER_SHEET = '服务器'


def _safe_sheet_name(name):
    """Excel sheet 名不允许的字符替换为下划线，且不超过 31 字符"""
    name = re.sub(r'[\[\]:*?/\\]', '_', str(name))
    return name[:31]


def _port_row(dev, p, conn_by_port):
    """一行端口数据（Port_Design 行，含设备名/位置列）"""
    disp = port_display_name(p)
    loc = dev['location'] or ''
    conn = conn_by_port.get(p['id'])
    if conn:
        # 本端各字段：取连接中属于本设备那端的值（各端写各端）
        if conn['port_a_id'] == p['id']:
            agg = conn['aggregation_group_a'] or ''
            agg_mode = conn['aggregation_mode_a'] or ''
            if_mode = conn['interface_mode_a'] or ''
            vlan = conn['vlan_id_a'] or ''
            desc, note = conn['description_a'] or '', conn['note_a'] or ''
            remote_dev, remote_port = conn['device_b_name'], conn_port_b_display(conn)
        else:
            agg = conn['aggregation_group_b'] or ''
            agg_mode = conn['aggregation_mode_b'] or ''
            if_mode = conn['interface_mode_b'] or ''
            vlan = conn['vlan_id_b'] or ''
            desc, note = conn['description_b'] or '', conn['note_b'] or ''
            remote_dev, remote_port = conn['device_a_name'], conn_port_a_display(conn)
        return [dev['name'], loc, disp, p['interface_status'] or '', agg,
                agg_mode, if_mode, vlan, remote_dev, remote_port, desc, note]
    return [dev['name'], loc, disp, p['interface_status'] or '', '', '', '', '', '', '', '', '']


def _write_device_sheet(wb, dev, conn_by_port):
    """单台设备的 Port_Design sheet"""
    model = dev['model'] or ''
    ws = wb.create_sheet(_safe_sheet_name(dev['name']))
    ws.append([f"{dev['name']} - Port_Design {model}".rstrip()])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append(HEADERS)
    for cell in ws[2]:
        cell.font = Font(bold=True)
    for p in get_device_ports(dev['id']):
        ws.append(_port_row(dev, p, conn_by_port))
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_project(project_id):
    """导出项目为 Port_Design Excel，返回 BytesIO。
    服务器模板设备自动汇总到一张「服务器」sheet（用设备列区分）；
    自由设备（无模板）不生成独立 sheet，只出现在对端列。
    """
    devices = get_project_devices(project_id)
    conns = get_connections(project_id)
    # 端口 id -> 连接（一条链路两端都可索引）
    conn_by_port = {}
    for c in conns:
        conn_by_port[c['port_a_id']] = c
        conn_by_port[c['port_b_id']] = c

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认空白 sheet

    servers = []
    for dev in devices:
        if dev['template_id'] is None:
            continue  # 自由设备不生成 sheet，只出现在对端列
        if dev.get('template_type') == 'server':
            servers.append(dev)
            continue
        _write_device_sheet(wb, dev, conn_by_port)

    # ---- 服务器汇总 sheet（一台服务器都不再单独成表） ----
    if servers:
        ws = wb.create_sheet(SERVER_SHEET)
        ws.append([f'{SERVER_SHEET} - Port_Design'])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append(HEADERS)
        for cell in ws[2]:
            cell.font = Font(bold=True)
        for dev in servers:
            for p in get_device_ports(dev['id']):
                ws.append(_port_row(dev, p, conn_by_port))
        for i, w in enumerate(WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # 项目无任何设备时，至少生成一个提示 sheet（openpyxl 不允许保存无 sheet 的工作簿）
    if not wb.sheetnames:
        ws = wb.create_sheet('项目为空')
        ws.append(['该项目暂无设备，无法导出端口数据'])
        ws['A1'].font = Font(bold=True)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def export_filename(project_id):
    """导出文件名 = 项目名_Port_Design.xlsx"""
    project = get_project(project_id)
    return f"{project['name']}_Port_Design.xlsx" if project else 'Port_Design.xlsx'


# ==================== 设备标签 / 线缆标签导出 ====================

# 设备标签可勾选列（列名 -> 设备行字段）
DEVICE_LABEL_COLUMNS = {
    '设备名称': 'name',
    '设备位置': 'location',
    '设备型号': 'model',
    '管理IP': 'management_ip',
    'BMC IP': 'bmc_ip',
}
DEFAULT_CABLE_FORMAT = '{位置}-{设备}-{接口}'


def render_label(fmt, **vals):
    """标签格式串渲染：{占位符} 替换为实际值，空值占位符连同其后分隔符一并删除。

    {位置}-{设备}-{接口} 且位置为空 -> "SW1-Fi_1"（不留横线）；
    无占位符原样返回；全部为空返回空串。
    """
    tokens = list(re.finditer(r'\{(\w+)\}', fmt))
    if not tokens:
        return fmt
    out = ''
    last_sep = fmt[:tokens[0].start()]  # 前导 literal，给第一个非空段
    for i, m in enumerate(tokens):
        val = str(vals.get(m.group(1), '') or '')
        if val:
            out += last_sep + val
            end = tokens[i + 1].start() if i + 1 < len(tokens) else len(fmt)
            last_sep = fmt[m.end():end]
        # 空值：跳过（分隔符随空段丢弃，下一个非空段仍用上一个非空段的分隔符）
    return out + (fmt[tokens[-1].end():] if out else '')


def _label_dev_row(dev, columns):
    """设备标签一行：勾选列 -> 值（空值留空）"""
    return [dev[DEVICE_LABEL_COLUMNS[c]] or '' for c in columns]


def export_device_labels(project_id, columns):
    """设备标签 xlsx：全部设备（含未连接），列按传入顺序（勾选列的有序列表）。"""
    devices = get_project_devices(project_id)
    # 保持传入顺序过滤（列顺序由弹窗设置决定）
    columns = [c for c in columns if c in DEVICE_LABEL_COLUMNS] or list(DEVICE_LABEL_COLUMNS)
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('设备标签')
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for dev in devices:
        ws.append(_label_dev_row(dev, columns))
    for i, w in enumerate([14, 12, 18, 15, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _cable_label_row(dev, p, conn, fmt, dev_info):
    """一行标签：本端设备（dev/端口 p）在 From、对端在 To（对端位置从 dev_info 查）"""
    if conn['port_a_id'] == p['id']:
        remote_id, remote_name, remote_port = conn['device_b_id'], conn['device_b_name'], conn['port_b_name']
    else:
        remote_id, remote_name, remote_port = conn['device_a_id'], conn['device_a_name'], conn['port_a_name']
    rdev = dev_info.get(remote_id)
    return [
        render_label(fmt, 位置=dev['location'] or '', 设备=dev['name'], 接口=p['port_name']),
        render_label(fmt, 位置=(rdev['location'] if rdev else '') or '',
                     设备=remote_name, 接口=remote_port),
    ]


def cable_label_preview(project_id, fmt, limit=6):
    """线缆标签预览：前 limit 行（与导出同源：按设备遍历，本端在 From）"""
    dev_info = {d['id']: d for d in get_project_devices(project_id)}
    rows = []
    for dev in get_project_devices(project_id):
        for p in get_device_ports(dev['id']):
            conn = get_connection_by_port(p['id'])
            if conn:
                rows.append(_cable_label_row(dev, p, conn, fmt, dev_info))
                if len(rows) >= limit:
                    return rows
    return rows


def export_cable_labels(project_id, fmt):
    """线缆标签 xlsx：单 sheet 两列 From|To。
    按设备遍历（设备顺序 + 端口顺序）：每行 = 该设备一个连接端口（本端在 From、对端在 To）；
    每根线缆在两端设备各出现一次——线缆两端标签都补齐。"""
    devices = get_project_devices(project_id)
    dev_info = {d['id']: d for d in devices}
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('线缆标签')
    ws.append(['From', 'To'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for dev in devices:
        for p in get_device_ports(dev['id']):
            conn = get_connection_by_port(p['id'])
            if conn:
                ws.append(_cable_label_row(dev, p, conn, fmt, dev_info))
    for i, w in enumerate([42, 42], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def device_label_filename(project_id):
    project = get_project(project_id)
    return f"{project['name']}_设备标签.xlsx" if project else '设备标签.xlsx'


def cable_label_filename(project_id):
    project = get_project(project_id)
    return f"{project['name']}_线缆标签.xlsx" if project else '线缆标签.xlsx'
