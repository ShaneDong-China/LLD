# -*- coding: utf-8 -*-
"""模板库 Excel 交换：导出 / 导入（简单格式 v2）

模板库文件标准：
- Sheet「模板信息」: 型号 | 类型 | 厂商 | 描述（类型以信息表为准，不做推断）
- 每个模板一个 sheet，sheet 名 = 型号，第一行表头「接口」，
  之后每行一个端口名；端口名即完整接口名，槽位/速率/介质/类型全部
  体现在名字里（如 Slot2_Fi_10GE_41、Cu_MGMT、NIC01），原样记录。

规则只有一条：sheet 名 = 型号。
"""
from contextlib import closing
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .db_utils import get_conn, query_all
from .template_utils import list_templates, list_ports, type_label
from .export_utils import _safe_sheet_name
from .import_utils import guess_port_attrs

INFO_SHEET = '模板信息'
INFO_HEADERS = ['型号', '类型', '厂商', '描述']
SHEET_HEADERS = ['接口']
# 模板库为空时的内置示例（S6805-56HF-G），导出供模仿填写
_SEED_SAMPLE = {
    'model': 'S6805-56HF-G', 'type': 'network', 'vendor': 'H3C',
    'description': '示例模板（模板库为空时的导出样例）',
    'ports': (['Cu_MGMT']
              + [f'Fi_10GE_{i:02d}' for i in range(1, 49)]
              + [f'Fi_100GE_{i}' for i in range(49, 57)]),
}
TYPE_MAP = {'网络设备': 'network', '服务器': 'server', '其它设备': 'other'}
# 表头列名的宽松匹配（中文为准，也认常见英文写法）
COL_KEYS = {
    '接口': ['接口', 'Interface'],
}


def export_templates():
    """导出全部模板为简单格式 Excel，返回 BytesIO"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认空白 sheet

    templates = list_templates()

    # ---- 「模板信息」sheet（元数据，导入时可选读取） ----
    # 模板库为空时，导出内置示例（S6805-56HF-G），方便模仿填写数据
    sources = templates if templates else [_SEED_SAMPLE]
    ws_info = wb.create_sheet(INFO_SHEET)
    ws_info.append(INFO_HEADERS)
    for cell in ws_info[1]:
        cell.font = Font(bold=True)
    for t in sources:
        ws_info.append([t['model'], type_label(t['type']), t['vendor'] or '', t['description'] or ''])
    for i, w in enumerate([26, 12, 16, 40], start=1):
        ws_info.column_dimensions[get_column_letter(i)].width = w

    # ---- 每模板一个 sheet：表头 + 每行一个端口 ----
    used_sheet_names = set()
    for t in sources:
        name = _safe_sheet_name(t['model'])
        while name in used_sheet_names:
            name = _safe_sheet_name(f"{t['model']}_{len(used_sheet_names) + 1}")
        used_sheet_names.add(name)
        ws = wb.create_sheet(name)

        ws.append(SHEET_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ports = t['ports'] if 'ports' in t else [p['port_name'] for p in list_ports(t['id'])]
        for pname in ports:
            ws.append([pname])  # 端口名即完整接口名（Slot2_Fi_10GE_41）
        ws.column_dimensions['A'].width = 28

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def export_filename():
    """模板库导出文件名"""
    return '模板库.xlsx'


def _read_info_sheet(source):
    """读取「模板信息」sheet 元数据: {型号: {type, vendor, description}}，无该 sheet 返回空 dict"""
    meta = {}
    try:
        wb = load_workbook(source, read_only=True, data_only=True)
    except Exception:
        return meta
    try:
        if INFO_SHEET not in wb.sheetnames:
            return meta
        rows = list(wb[INFO_SHEET].iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return meta
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    try:
        i_model, i_type = header.index('型号'), header.index('类型')
    except ValueError:
        return meta
    i_vendor = header.index('厂商') if '厂商' in header else -1
    i_desc = header.index('描述') if '描述' in header else -1

    def val(row, idx):
        if idx < 0 or idx >= len(row) or row[idx] is None:
            return ''
        return str(row[idx]).strip()

    for row in rows[1:]:
        model = val(row, i_model)
        if not model:
            continue
        type_ = TYPE_MAP.get(val(row, i_type), '')
        meta[model] = {'type': type_, 'vendor': val(row, i_vendor),
                       'description': val(row, i_desc)}
    return meta


def _parse_library_sheets(source):
    """解析模板库端口 sheet（简单格式：仅「接口」列，端口名原样记录）。
    返回 (templates: {型号: [端口名]}, skipped: [{'name', 'reason'}])
    """
    templates = {}
    skipped = []
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            if ws.title == INFO_SHEET:
                continue
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                skipped.append({'name': ws.title, 'reason': '空 sheet（无数据）'})
                continue
            # 找表头行（前 5 行内含「接口」列）
            header_row = None
            for i, row in enumerate(rows[:5]):
                cells = [str(c).strip() if c is not None else '' for c in row]
                if any(any(k in c for k in COL_KEYS['接口']) for c in cells):
                    header_row = i
                    break
            if header_row is None:
                skipped.append({'name': ws.title, 'reason': '未找到「接口」表头'})
                continue

            def col_idx(keys):
                for j, c in enumerate(rows[header_row]):
                    s = str(c).strip() if c is not None else ''
                    if any(k == s or (k and k in s) for k in keys):
                        return j
                return None

            col_iface = col_idx(COL_KEYS['接口'])
            if col_iface is None:
                skipped.append({'name': ws.title, 'reason': '未找到「接口」表头'})
                continue

            ports = []
            for row in rows[header_row + 1:]:
                if col_iface >= len(row) or row[col_iface] is None:
                    continue
                pname = str(row[col_iface]).strip()
                if pname:
                    ports.append(pname)
            templates[ws.title] = ports
    finally:
        wb.close()
    return templates, skipped


def classify_templates(source):
    """解析并分类模板库文件（预览与导入共用同一分类标准，保证显示与实际一致）。
    返回 {'importable': [{'model', 'ports', 'vendor', 'type', 'description'}],
          'skipped': [{'name', 'reason'}]}
    跳过原因（严格模式）：
    - 库中已存在（不覆盖）
    - 模板信息表中未定义 / 类型无效
    - 格式问题（空 sheet / 未找到「接口」表头）
    - 模板信息表已定义但文件无对应 sheet
    """
    if hasattr(source, 'seek'):
        source.seek(0)
    meta = _read_info_sheet(source)
    if hasattr(source, 'seek'):
        source.seek(0)
    templates, fmt_skipped = _parse_library_sheets(source)
    existing = {r['model'] for r in query_all('SELECT model FROM device_templates')}

    importable, skipped = [], list(fmt_skipped)
    for model, ports in templates.items():
        if not model:
            skipped.append({'name': '（无型号 sheet）', 'reason': 'sheet 名为空'})
            continue
        if model in existing:
            skipped.append({'name': model, 'reason': '库中已存在（不覆盖）'})
            continue
        info = meta.get(model)
        if not info:
            skipped.append({'name': model, 'reason': '模板信息表中未定义'})
            continue
        if not info.get('type'):
            skipped.append({'name': model, 'reason': '模板信息表中类型无效（需为 网络设备/服务器/其它设备）'})
            continue
        importable.append({
            'model': model, 'ports': ports,
            'vendor': info.get('vendor', ''),
            'type': info.get('type', ''),
            'description': info.get('description', ''),
        })
    # 信息表已定义但文件里没有对应 sheet 的型号（文件可能漏了该 sheet）
    for model in meta:
        if model not in templates:
            skipped.append({'name': model, 'reason': '模板信息表已定义但文件无对应 sheet'})
    return {'importable': importable, 'skipped': skipped}


def import_templates(source):
    """从模板库 Excel 导入模板（严格模式，单事务，失败自动回滚）。
    分类标准与 classify_templates 完全一致：只导入可导入项，其余跳过。返回摘要 dict。"""
    result = classify_templates(source)
    created = []
    port_total = 0
    with closing(get_conn()) as conn, conn:
        for t in result['importable']:
            cur = conn.execute(
                'INSERT INTO device_templates (model, vendor, type, description) VALUES (?,?,?,?)',
                (t['model'], t['vendor'], t['type'], t['description']))
            tpl_id = cur.lastrowid
            # 端口：端口名即完整接口名（Slot2_Fi_10GE_41 原样记录）
            for port_name in t['ports']:
                conn.execute(
                    'INSERT INTO port_templates (device_template_id, port_name) VALUES (?,?)',
                    (tpl_id, port_name))
                port_total += 1
            created.append(t['model'])
    return {'created': created, 'skipped': result['skipped'], 'port_count': port_total}
