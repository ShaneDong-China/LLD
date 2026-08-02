# -*- coding: utf-8 -*-
"""Excel 导入：解析 Port_Design 格式工作簿，迁移为项目数据

两阶段设计：
1. 解析阶段：扫描全部 sheet，汇总"设备 -> 端口"计划与连接行（跨 sheet 引用也能正确解析）
2. 导入阶段：在单个事务内建模板/设备/端口，再建连接（按链路对去重，两端聚合组合并）

默认跳过非 Port_Design 格式的 sheet（如"服务器端口信息"）。
"""
import os
import re
from contextlib import closing

from openpyxl import load_workbook

from .db_utils import get_conn

TITLE_MARK = 'Port_Design'


def _clean(v):
    """单元格值清洗：去空白，'/' 视为空"""
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s == '/' else s


def guess_port_attrs(name):
    """根据端口名猜测类型/速率/介质"""
    n = name.upper()
    if n.startswith('CU_') or 'MGMT' in n or n.startswith('BMC') or n.startswith('IBMC'):
        ptype = '电口'
        media = 'RJ45'
        # 管理口/电口默认 1GE（名字里含 GE 或 MGMT/BMC 管理口）
        speed = '10GE' if '10GE' in n else ('25GE' if '25GE' in n else ('40GE' if '40GE' in n
                else ('100GE' if '100GE' in n else ('1GE' if ('GE' in n or 'MGMT' in n or 'BMC' in n) else ''))))
    else:
        ptype = '光口'
        speed = '100GE' if '100GE' in n else ('40GE' if '40GE' in n else ('25GE' if '25GE' in n
                else ('10GE' if '10GE' in n else '')))
        media = {'100GE': 'QSFP28', '40GE': 'QSFP+', '25GE': 'SFP28', '10GE': 'SFP+'}.get(speed, '')
    return ptype, speed, media


def _strip_slot_prefix(name):
    """去掉端口名里的槽位前缀（Slot2_Fi_10GE_41 -> Fi_10GE_41）"""
    return re.sub(r'^Slot\d+_', '', name)


def parse_workbook(source, require_title=True):
    """解析工作簿（路径或文件对象），返回全局计划。
    require_title=True（项目导入）要求标题行含 Port_Design 标记；
    False（模板库导入）时允许无标题行，靠表头识别，型号留空。
    计划结构: {devices: {设备名: {model, ports: {端口名: {slot, status}}}},
               rows: [{device, port, status, agg, agg_mode, if_mode, vlan,
                       remote_device, remote_port, description, note}],
               skipped: [sheet名], warnings: [str]}
    """
    plan = {'devices': {}, 'rows': [], 'skipped': [], 'warnings': []}
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            try:
                _parse_sheet(ws, plan, require_title)
            except Exception as e:  # 单个 sheet 解析失败不影响其他
                plan['skipped'].append(f"{ws.title}（解析失败: {e}）")
    finally:
        wb.close()
    return plan


def _parse_sheet(ws, plan, require_title=True):
    """解析单个 Port_Design sheet"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        plan['skipped'].append(ws.title)
        return

    # 找标题行（含 Port_Design 的行），从中提取型号
    title_row = None
    for i, row in enumerate(rows[:15]):
        if any(isinstance(c, str) and TITLE_MARK in c for c in row):
            title_row = i
            break
    model = ''
    if title_row is not None:
        for c in rows[title_row]:
            if isinstance(c, str) and TITLE_MARK in c:
                model = c.split(TITLE_MARK, 1)[1].strip()
                break
    elif require_title:
        plan['skipped'].append(ws.title)
        return

    # 找子表头行（含 Interface Status），其下一行为数据起点
    header_row = None
    search_start = title_row if title_row is not None else 0
    for i in range(search_start, min(search_start + 10, len(rows))):
        cells = [str(c).strip() if c is not None else '' for c in rows[i]]
        if 'Interface Status' in cells or '接口状态' in cells:
            header_row = i
            break
    if header_row is None:
        plan['skipped'].append(f"{ws.title}（未找到表头）")
        return

    header = rows[header_row]

    def col_index(name):
        for j, c in enumerate(header):
            if c is not None and str(c).strip() == name:
                return j
        return None

    col_dev, col_slot, col_iface = col_index('Device Name'), col_index('Device Slot'), col_index('Interface')
    col_status, col_agg = col_index('Interface Status'), col_index('Aggregation Group')
    col_agg_mode, col_if_mode = col_index('Aggregation Mode'), col_index('Interface Mode')
    col_vlan = col_index('Vlan ID')
    # 描述/备注表头在子表头上方一行（合并单元格），需在两行中查找
    desc_row = header_row if col_index('Connection Description') is not None else header_row - 1
    note_row = header_row if col_index('Note') is not None else header_row - 1
    if desc_row >= 0:
        for j, c in enumerate(rows[desc_row]):
            if c is not None and str(c).strip() == 'Connection Description':
                col_desc = j
                break
    else:
        col_desc = None
    if note_row >= 0:
        for j, c in enumerate(rows[note_row]):
            if c is not None and str(c).strip() == 'Note':
                col_note = j
                break
    else:
        col_note = None
    # 远程表头：第二次出现的 Device Name / Interface
    col_rdev, col_rport = None, None
    found_dev = found_iface = 0
    for j, c in enumerate(header):
        if c is not None and str(c).strip() == 'Device Name':
            found_dev += 1
            if found_dev == 2:
                col_rdev = j
        if c is not None and str(c).strip() == 'Interface':
            found_iface += 1
            if found_iface == 2:
                col_rport = j

    def cell(row, col):
        if col is None or col >= len(row):
            return ''
        return _clean(row[col])

    # 读取数据行（forward-fill 设备名/槽位/状态）
    cur_dev = cur_slot = cur_status = ''
    any_data = False
    for row in rows[header_row + 1:]:
        port = cell(row, col_iface)
        if not port:
            continue
        dev = cell(row, col_dev)
        if dev:
            cur_dev = dev
        if not cur_dev:
            continue
        slot = cell(row, col_slot)
        if slot:
            cur_slot = slot
        status = cell(row, col_status)
        if status:
            cur_status = status
        any_data = True

        # 登记设备与端口
        dev_plan = plan['devices'].setdefault(cur_dev, {'model': model, 'ports': {}})
        dev_plan['ports'].setdefault(port, {'slot': cur_slot, 'status': cur_status})

        # 登记连接行
        rdev = cell(row, col_rdev)
        if rdev:
            plan['rows'].append({
                'device': cur_dev, 'port': port, 'status': cur_status,
                'agg': cell(row, col_agg),
                'agg_mode': cell(row, col_agg_mode),
                'if_mode': cell(row, col_if_mode),
                'vlan': cell(row, col_vlan),
                'remote_device': rdev,
                'remote_port': cell(row, col_rport),
                'description': cell(row, col_desc),
                'note': cell(row, col_note),
            })
    if not any_data:
        plan['skipped'].append(f"{ws.title}（无数据行）")


def import_workbook(source, project_name='', include_other=False):
    """把工作簿导入为一个新项目，返回统计摘要 dict（单事务，失败自动回滚）"""
    plan = parse_workbook(source)
    warnings = list(plan['warnings'])
    if not plan['devices']:
        raise ValueError('未解析到任何设备数据，请确认文件为 Port_Design 格式')

    if not project_name:
        if isinstance(source, str):
            base = source
        else:
            base = getattr(source, 'name', '导入项目') or '导入项目'
        project_name = os.path.splitext(os.path.basename(base))[0]

    new_templates, dict_added, vlan_added, free_devices = [], [], [], set()
    conn_created = conn_merged = 0

    def uniq_project_name(name):
        """项目名重名自动加后缀"""
        if not conn.execute('SELECT id FROM projects WHERE name = ?', (name,)).fetchone():
            return name
        i = 2
        while conn.execute('SELECT id FROM projects WHERE name = ?', (f'{name} ({i})',)).fetchone():
            i += 1
        return f'{name} ({i})'

    def uniq_template_id(model, ports):
        """按型号取模板 id，不存在则创建（type 由是否有槽位推断）"""
        row = conn.execute('SELECT id FROM device_templates WHERE model = ?', (model,)).fetchone()
        if row:
            return row['id']
        type_ = 'network'
        cur = conn.execute(
            'INSERT INTO device_templates (model, vendor, type, description) VALUES (?,?,?,?)',
            (model, '', type_, '由导入自动创建（可修改）'))
        new_templates.append(model)
        return cur.lastrowid

    def resolve_port(dev_name, port_name, is_remote):
        """解析设备端口，不存在则按需创建（自由设备/自由端口），返回 port_id"""
        row = conn.execute(
            'SELECT id, template_id FROM project_devices WHERE project_id = ? AND name = ?',
            (project_id, dev_name)).fetchone()
        if row:
            dev_id = row['id']
            p = conn.execute(
                'SELECT id FROM project_ports WHERE project_device_id = ? AND port_name = ?',
                (dev_id, port_name)).fetchone()
            if p:
                return p['id']
            cur = conn.execute(
                'INSERT INTO project_ports (project_device_id, port_name) VALUES (?,?)',
                (dev_id, port_name))
            # 只有有模板的设备上找不到端口才可能是笔误，值得警告；自由设备按需建端口是常态
            if row['template_id'] is not None:
                warnings.append(f'设备 {dev_name} 上未找到端口 {port_name}（已按自由端口创建，请检查拼写）')
            return cur.lastrowid
        # 未知设备：创建自由设备 + 自由端口
        cur = conn.execute(
            'INSERT INTO project_devices (project_id, template_id, name) VALUES (?,?,?)',
            (project_id, None, dev_name))
        dev_id = cur.lastrowid
        free_devices.add(dev_name)
        cur = conn.execute(
            'INSERT INTO project_ports (project_device_id, port_name) VALUES (?,?)',
            (dev_id, port_name))
        if is_remote:
            warnings.append(f'对端设备 {dev_name} 不在本项目设备中，已自动创建（自由设备）')
        return cur.lastrowid

    with closing(get_conn()) as conn, conn:
        # ---- 阶段1: 模板 / 项目 / 设备 / 端口 ----
        final_name = uniq_project_name(project_name)
        cur = conn.execute('INSERT INTO projects (name, description) VALUES (?,?)',
                           (final_name, '由 Excel 导入'))
        project_id = cur.lastrowid

        for dev_name, dev_plan in plan['devices'].items():
            model = dev_plan['model'] or ''
            tpl_id = uniq_template_id(model, dev_plan['ports']) if model else None
            cur = conn.execute(
                'INSERT INTO project_devices (project_id, template_id, name) VALUES (?,?,?)',
                (project_id, tpl_id, dev_name))
            dev_id = cur.lastrowid
            for port_name, info in dev_plan['ports'].items():
                # 模板端口不存在则自动扩充模板库
                template_port_id = None
                if tpl_id is not None:
                    t = conn.execute(
                        'SELECT id FROM port_templates WHERE device_template_id=? AND port_name=?',
                        (tpl_id, port_name)).fetchone()
                    if t:
                        template_port_id = t['id']
                    else:
                        slot_id = None
                        if info['slot']:
                            s = conn.execute(
                                'SELECT id FROM slot_templates WHERE device_template_id=? AND slot_number=?',
                                (tpl_id, info['slot'])).fetchone()
                            if s:
                                slot_id = s['id']
                            else:
                                cur2 = conn.execute(
                                    'INSERT INTO slot_templates (device_template_id, slot_number, slot_type, description) '
                                    'VALUES (?,?,?,?)',
                                    (tpl_id, info['slot'], '业务板', '由导入自动创建'))
                                slot_id = cur2.lastrowid
                        ptype, speed, media = guess_port_attrs(port_name)
                        cur2 = conn.execute(
                            'INSERT INTO port_templates (device_template_id, slot_template_id, port_name, port_type, speed, media_type) '
                            'VALUES (?,?,?,?,?,?)',
                            (tpl_id, slot_id, port_name, ptype, speed, media))
                        template_port_id = cur2.lastrowid
                conn.execute(
                    'INSERT INTO project_ports (project_device_id, template_port_id, port_name, interface_status) '
                    'VALUES (?,?,?,?)',
                    (dev_id, template_port_id, port_name, info['status'] or 'Down'))

        # ---- 阶段2: 连接（按链路对去重，镜像行合并）----
        link_map = {}  # (port_a, port_b) -> connection_id
        for r in plan['rows']:
            local_id = resolve_port(r['device'], r['port'], is_remote=False)
            remote_id = resolve_port(r['remote_device'], r['remote_port'], is_remote=True)
            port_a, port_b = sorted((local_id, remote_id))

            # 聚合组归属：本行是本端视角，聚合组写到本端
            agg_a, agg_b = (r['agg'], '') if port_a == local_id else ('', r['agg'])

            # 字典 / VLAN 自动扩充
            for cat, val in (('aggregation_mode', r['agg_mode']), ('interface_mode', r['if_mode'])):
                if val:
                    row = conn.execute('SELECT id FROM option_dicts WHERE category=? AND value=?', (cat, val)).fetchone()
                    if not row:
                        conn.execute('INSERT INTO option_dicts (category, value, sort_order) VALUES (?,?,'
                                     '(SELECT COALESCE(MAX(sort_order),0)+1 FROM option_dicts WHERE category=?))',
                                     (cat, val, cat))
                        dict_added.append(f"{cat}:{val}")
            if r['status'] and r['status'] not in ('Down', 'UP'):
                row = conn.execute('SELECT id FROM option_dicts WHERE category=? AND value=?',
                                   ('interface_status', r['status'])).fetchone()
                if not row:
                    conn.execute('INSERT INTO option_dicts (category, value, sort_order) VALUES (?,?,'
                                 '(SELECT COALESCE(MAX(sort_order),0)+1 FROM option_dicts WHERE category=?))',
                                 ('interface_status', r['status'], 'interface_status'))
                    dict_added.append(f"interface_status:{r['status']}")
            for v in re.split(r'[\s,、]+', r['vlan']):
                if v:
                    row = conn.execute('SELECT id FROM project_vlans WHERE project_id=? AND value=?',
                                       (project_id, v)).fetchone()
                    if not row:
                        conn.execute('INSERT INTO project_vlans (project_id, value, sort_order) VALUES (?,?,'
                                     '(SELECT COALESCE(MAX(sort_order),0)+1 FROM project_vlans WHERE project_id=?))',
                                     (project_id, v, project_id))
                        vlan_added.append(v)

            if (port_a, port_b) in link_map:
                # 镜像行：补齐对端聚合组与空缺字段
                cid = link_map[(port_a, port_b)]
                cur = conn.execute('SELECT * FROM project_connections WHERE id = ?', (cid,)).fetchone()
                sets, params = [], []
                for col, val in [('aggregation_group_a', agg_a), ('aggregation_group_b', agg_b),
                                 ('aggregation_mode', r['agg_mode']), ('interface_mode', r['if_mode']),
                                 ('vlan_id', r['vlan']), ('description', r['description']), ('note', r['note'])]:
                    if val and not cur[col]:
                        sets.append(f'{col} = ?')
                        params.append(val)
                if sets:
                    conn.execute(f'UPDATE project_connections SET {", ".join(sets)} WHERE id = ?', params + [cid])
                conn_merged += 1
                continue

            cur = conn.execute('''
                INSERT INTO project_connections
                    (project_id, port_a_id, port_b_id, aggregation_group_a, aggregation_group_b,
                     aggregation_mode, interface_mode, vlan_id, description, note)
                VALUES (?,?,?,?,?,?,?,?,?,?)''',
                (project_id, port_a, port_b, agg_a or None, agg_b or None,
                 r['agg_mode'] or None, r['if_mode'] or None, r['vlan'] or None,
                 r['description'] or None, r['note'] or None))
            link_map[(port_a, port_b)] = cur.lastrowid
            conn_created += 1

        # ---- 阶段3: 依据连接重算端口占用 ----
        conn.execute('''
            UPDATE project_ports SET is_used =
                EXISTS(SELECT 1 FROM project_connections c
                       WHERE c.port_a_id = project_ports.id OR c.port_b_id = project_ports.id)
            WHERE project_device_id IN (SELECT id FROM project_devices WHERE project_id = ?)''',
            (project_id,))

        # 端口被多条连接引用（源表脏数据遗留，如实提示）
        conflicts = conn.execute('''
            SELECT COUNT(*) AS n FROM (
                SELECT port_id FROM (
                    SELECT port_a_id AS port_id FROM project_connections
                    UNION ALL SELECT port_b_id FROM project_connections)
                GROUP BY port_id HAVING COUNT(*) > 1)''').fetchone()['n']
        if conflicts:
            warnings.append(f'有 {conflicts} 个端口被多条连接同时引用（源表中该端口存在重复连接，请核实后删除多余链路）')

    return {
        'project_id': project_id,
        'project_name': final_name,
        'device_count': len(plan['devices']) + len(free_devices - set(plan['devices'])),
        'free_device_count': len(free_devices),
        'connection_count': conn_created,
        'connection_merged': conn_merged,
        'new_templates': new_templates,
        'dict_added': dict_added,
        'vlan_added': vlan_added,
        'skipped': plan['skipped'],
        'warnings': warnings,
    }
